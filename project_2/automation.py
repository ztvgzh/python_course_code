import openpyxl as o
"""import os  #to understand where code is running
print(os.getcwd())"""

inv_file = o.load_workbook("inventory.xlsx")
product_list = inv_file['Sheet1']

profucts_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2, product_list.max_row+1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = int(product_list.cell(product_row, 1).value)
    inventory_price = product_list.cell(product_row, 5)

    #task 1. to find the quantity of products per supplier
    if supplier_name in profucts_per_supplier:
        current_num_products = profucts_per_supplier.get(supplier_name)
        profucts_per_supplier[supplier_name] = current_num_products+1
    else:
        profucts_per_supplier[supplier_name] = 1
# task 1. output: print(profucts_per_supplier)

    #task 2. to find total value of products for each supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory*price
# task 2. output: print(total_value_per_supplier)

    #task 3. to find inventory that less 10
    if inventory<10:
        products_under_10_inv[product_num] = int(inventory)
# task 3. output: print(products_under_10_inv)

    # task 4. new column
    inventory_price.value = inventory*price
    
inv_file.save("inventory_with_total_price.xlsx")